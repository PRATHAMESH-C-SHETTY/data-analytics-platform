from datetime import datetime
import os


# ══════════════════════════════════════════════════════════════
#  PDF REPORT
# ══════════════════════════════════════════════════════════════
def create_pdf_report(path, dataset, df, stats, charts, insights):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                     TableStyle, HRFlowable, PageBreak, KeepTogether)

    INDIGO  = colors.HexColor('#6366f1')
    DARK    = colors.HexColor('#0f172a')
    SLATE   = colors.HexColor('#1e293b')
    MUTED   = colors.HexColor('#64748b')
    LIGHT   = colors.HexColor('#f8fafc')
    WHITE   = colors.white
    SUCCESS = colors.HexColor('#10b981')
    WARN    = colors.HexColor('#f59e0b')
    DANGER  = colors.HexColor('#ef4444')
    CYAN    = colors.HexColor('#06b6d4')

    SEV_COLORS = {'info': INDIGO, 'success': SUCCESS, 'warning': WARN, 'danger': DANGER}

    doc = SimpleDocTemplate(path, pagesize=A4,
                            leftMargin=1.8*cm, rightMargin=1.8*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    S = lambda name, **kw: ParagraphStyle(name, parent=styles['Normal'], **kw)
    title_s   = S('T',  fontSize=22, textColor=INDIGO, fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=4)
    sub_s     = S('Su', fontSize=11, textColor=MUTED,  fontName='Helvetica', alignment=TA_CENTER, spaceAfter=2)
    h2_s      = S('H2', fontSize=13, textColor=INDIGO, fontName='Helvetica-Bold', spaceBefore=14, spaceAfter=6)
    h3_s      = S('H3', fontSize=10, textColor=colors.HexColor('#334155'), fontName='Helvetica-Bold', spaceBefore=8, spaceAfter=4)
    body_s    = S('B',  fontSize=9,  textColor=colors.HexColor('#334155'), leading=14, spaceAfter=3)
    small_s   = S('Sm', fontSize=8,  textColor=MUTED, leading=12)
    center_s  = S('C',  fontSize=9,  textColor=MUTED, alignment=TA_CENTER)

    def styled_table(data, col_widths=None, header_bg=INDIGO):
        t = Table(data, colWidths=col_widths, repeatRows=1)
        row_bgs = [LIGHT if i % 2 == 0 else WHITE for i in range(len(data))]
        style = TableStyle([
            ('BACKGROUND',  (0,0), (-1,0), header_bg),
            ('TEXTCOLOR',   (0,0), (-1,0), WHITE),
            ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',    (0,0), (-1,-1), 8),
            ('PADDING',     (0,0), (-1,-1), 5),
            ('GRID',        (0,0), (-1,-1), 0.4, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [LIGHT, WHITE]),
            ('VALIGN',      (0,0), (-1,-1), 'MIDDLE'),
        ])
        t.setStyle(style)
        return t

    story = []

    # ── Cover
    story += [Spacer(1, 0.5*cm),
              Paragraph('DataLens Analytics', title_s),
              Paragraph('Comprehensive Data Analysis Report', sub_s),
              Paragraph(f'Dataset: <b>{dataset.name}</b>', sub_s),
              Paragraph(f'Generated: {datetime.now().strftime("%B %d, %Y at %H:%M UTC")}', small_s),
              HRFlowable(width='100%', thickness=2, color=INDIGO, spaceAfter=10)]

    # ── 1. Dataset Summary
    story.append(Paragraph('1. Dataset Summary', h2_s))
    summary = [
        ['Property', 'Value'],
        ['Dataset Name', dataset.name],
        ['File Format', dataset.file_type.upper()],
        ['Total Rows', f'{dataset.num_rows:,}'],
        ['Total Columns', str(dataset.num_cols)],
        ['File Size', f'{dataset.file_size:.3f} MB'],
        ['Upload Date', dataset.uploaded_at.strftime('%Y-%m-%d %H:%M')],
        ['Numeric Columns', str(len(stats.get('numeric', {})))],
        ['Categorical Columns', str(len(stats.get('categorical', {})))],
        ['Datetime Columns', str(len(stats.get('datetime', {})))],
    ]
    story.append(styled_table(summary, col_widths=[6*cm, 10*cm]))
    story.append(Spacer(1, 0.4*cm))

    # ── 2. Numeric Statistics
    if stats.get('numeric'):
        story.append(Paragraph('2. Numeric Column Statistics', h2_s))
        header = ['Column', 'Mean', 'Median', 'Std Dev', 'Min', 'Max', 'Skewness', 'CV %']
        rows   = [header]
        for col, s in stats['numeric'].items():
            rows.append([col, f"{s['mean']:.3f}", f"{s['median']:.3f}", f"{s['std']:.3f}",
                         f"{s['min']:.3f}", f"{s['max']:.3f}", f"{s['skewness']:.3f}", f"{s['cv']:.1f}"])
        story.append(styled_table(rows))
        story.append(Spacer(1, 0.4*cm))

    # ── 3. Categorical Statistics
    if stats.get('categorical'):
        story.append(Paragraph('3. Categorical Column Statistics', h2_s))
        header = ['Column', 'Unique Values', 'Top Value', 'Top Count', 'Top %']
        rows   = [header]
        for col, s in stats['categorical'].items():
            rows.append([col, str(s['unique_count']), str(s['top_value']),
                         str(s['top_count']), f"{s['top_pct']:.1f}%"])
        story.append(styled_table(rows))
        story.append(Spacer(1, 0.4*cm))

    # ── 4. Correlations
    if stats.get('top_correlations'):
        story.append(Paragraph('4. Top Correlations', h2_s))
        header = ['Column 1', 'Column 2', 'Correlation (r)', 'Strength']
        rows   = [header]
        for p in stats['top_correlations']:
            r = p['r']
            strength = 'Very Strong' if abs(r) > 0.85 else ('Strong' if abs(r) > 0.6 else 'Moderate')
            rows.append([p['col1'], p['col2'], f"{r:.3f}", strength])
        story.append(styled_table(rows))
        story.append(Spacer(1, 0.4*cm))

    # ── 5. Insights
    story.append(Paragraph('5. Automated Insights', h2_s))
    for ins in insights:
        sev_color = SEV_COLORS.get(ins.get('severity', 'info'), INDIGO)
        story.append(KeepTogether([
            Paragraph(f'<font color="#{_hex(sev_color)}">{ins["icon"]} [{ins["category"]}]</font> <b>{ins["title"]}</b>', h3_s),
            Paragraph(ins['text'], body_s),
        ]))

    # ── 6. Data Sample
    story.append(Paragraph('6. Data Sample (First 15 Rows)', h2_s))
    display_cols = list(df.columns[:8])
    rows = [display_cols]
    for _, row in df.head(15).iterrows():
        rows.append([str(row[c])[:20] for c in display_cols])
    story.append(styled_table(rows))

    # ── Footer
    story += [Spacer(1, 0.8*cm),
              HRFlowable(width='100%', thickness=1, color=MUTED),
              Paragraph('DataLens Analytics Platform — Confidential Report', center_s)]

    doc.build(story)


def _hex(c):
    # reportlab color -> hex string
    try:
        return '{:02x}{:02x}{:02x}'.format(int(c.red*255), int(c.green*255), int(c.blue*255))
    except:
        return '6366f1'


# ══════════════════════════════════════════════════════════════
#  EXCEL REPORT
# ══════════════════════════════════════════════════════════════
def create_excel_report(path, dataset, df, stats, insights):
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint

    INDIGO_HEX  = '6366F1'
    LIGHT_HEX   = 'EEF2FF'
    SLATE_HEX   = '1E293B'
    WHITE_HEX   = 'FFFFFF'

    thin = Side(style='thin', color='E2E8F0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(cell, val, bg=INDIGO_HEX):
        cell.value = val
        cell.font  = Font(bold=True, color=WHITE_HEX, size=10)
        cell.fill  = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    def data_cell(cell, val, light=False):
        cell.value  = val
        cell.font   = Font(size=9, color='334155')
        cell.fill   = PatternFill('solid', fgColor=LIGHT_HEX if light else WHITE_HEX)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border

    wb = openpyxl.Workbook()

    # ── Sheet 1: Cover / Summary ──
    ws = wb.active
    ws.title = 'Summary'
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 34

    ws['A1'] = '⚡ DataLens Analytics Report'
    ws['A1'].font = Font(bold=True, size=18, color=INDIGO_HEX)
    ws['A2'] = f'Dataset: {dataset.name}'
    ws['A2'].font = Font(size=11, color='64748B')
    ws['A3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A3'].font = Font(size=10, color='94A3B8')

    ws.append([])
    props = [('Dataset Name', dataset.name), ('File Format', dataset.file_type.upper()),
             ('Total Rows', f'{dataset.num_rows:,}'), ('Total Columns', dataset.num_cols),
             ('File Size (MB)', dataset.file_size), ('Upload Date', dataset.uploaded_at.strftime('%Y-%m-%d')),
             ('Numeric Columns', len(stats.get('numeric', {}))),
             ('Categorical Columns', len(stats.get('categorical', {})))]
    r = 5
    for k, v in props:
        ws.cell(r, 1, k).font = Font(bold=True, size=9, color='475569')
        ws.cell(r, 2, v)
        r += 1

    # ── Sheet 2: Raw Data ──
    ws2 = wb.create_sheet('Data')
    ws2.sheet_view.showGridLines = False
    for ci, col in enumerate(df.columns, 1):
        c = ws2.cell(1, ci)
        hdr(c, col)
        ws2.column_dimensions[get_column_letter(ci)].width = max(12, min(30, len(str(col)) + 4))

    for ri, row in enumerate(df.head(1000).itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            v = float(val) if isinstance(val, (int, float)) and not isinstance(val, bool) else str(val)
            data_cell(ws2.cell(ri, ci), v, ri % 2 == 0)

    # ── Sheet 3: Numeric Statistics ──
    if stats.get('numeric'):
        ws3 = wb.create_sheet('Numeric Stats')
        ws3.sheet_view.showGridLines = False
        header = ['Column', 'Mean', 'Median', 'Std Dev', 'Min', 'Max', 'Q25', 'Q75', 'Skewness', 'CV %']
        for ci, h in enumerate(header, 1):
            hdr(ws3.cell(1, ci), h)
            ws3.column_dimensions[get_column_letter(ci)].width = 13
        ws3.column_dimensions['A'].width = 22
        for ri, (col, s) in enumerate(stats['numeric'].items(), 2):
            vals = [col, s['mean'], s['median'], s['std'], s['min'], s['max'],
                    s['q25'], s['q75'], s['skewness'], s['cv']]
            for ci, v in enumerate(vals, 1):
                data_cell(ws3.cell(ri, ci), round(v, 4) if isinstance(v, float) else v, ri % 2 == 0)

        # Embedded bar chart
        chart = BarChart()
        chart.title = 'Mean Values'
        chart.style = 10
        data_ref = Reference(ws3, min_col=2, max_col=2, min_row=1, max_row=1+len(stats['numeric']))
        cats_ref = Reference(ws3, min_col=1, min_row=2, max_row=1+len(stats['numeric']))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        ws3.add_chart(chart, f'L2')

    # ── Sheet 4: Categorical Statistics ──
    if stats.get('categorical'):
        ws4 = wb.create_sheet('Categorical Stats')
        ws4.sheet_view.showGridLines = False
        header = ['Column', 'Unique Values', 'Top Value', 'Top Count', 'Top %']
        for ci, h in enumerate(header, 1):
            hdr(ws4.cell(1, ci), h)
        ws4.column_dimensions['A'].width = 22
        ws4.column_dimensions['C'].width = 28
        for ri, (col, s) in enumerate(stats['categorical'].items(), 2):
            vals = [col, s['unique_count'], s['top_value'], s['top_count'], f"{s['top_pct']:.1f}%"]
            for ci, v in enumerate(vals, 1):
                data_cell(ws4.cell(ri, ci), v, ri % 2 == 0)

    # ── Sheet 5: Insights ──
    ws5 = wb.create_sheet('Insights')
    ws5.sheet_view.showGridLines = False
    ws5.column_dimensions['A'].width = 18
    ws5.column_dimensions['B'].width = 30
    ws5.column_dimensions['C'].width = 70
    hdr(ws5.cell(1,1), 'Category')
    hdr(ws5.cell(1,2), 'Title')
    hdr(ws5.cell(1,3), 'Insight')
    for ri, ins in enumerate(insights, 2):
        ws5.cell(ri,1).value = f"{ins['icon']} {ins['category']}"
        ws5.cell(ri,2).value = ins['title']
        ws5.cell(ri,3).value = ins['text']
        for ci in range(1, 4):
            c = ws5.cell(ri, ci)
            c.font = Font(size=9, color='334155')
            c.fill = PatternFill('solid', fgColor=LIGHT_HEX if ri%2==0 else WHITE_HEX)
            c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            c.border = border
        ws5.row_dimensions[ri].height = 40

    wb.save(path)
